from fastapi import FastAPI, Query, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel
from pathlib import Path
from typing import Optional, List
import pandas as pd
from openpyxl import load_workbook
import math

app = FastAPI()
FOLDER = Path(__file__).parent


def load_data() -> pd.DataFrame:
    frames = []
    for year in ["2023", "2024", "2025"]:
        path = FOLDER / f"{year}.xlsx"
        if path.exists():
            df = pd.read_excel(path)
            df["연도"] = year
            frames.append(df)
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True)
    for col in ["수량", "매출금액", "실매출금액", "실매입금액", "실이익금액", "실이익율", "기준가매출액", "기준가"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df.dropna(subset=["실매출금액"])


def filter_df(df, years, customers, manufacturers, suppliers, reps, min_s, max_s):
    if years:
        df = df[df["연도"].isin(years)]
    if customers:
        df = df[df["매출처"].isin(customers)]
    if manufacturers:
        df = df[df["제조사"].isin(manufacturers)]
    if suppliers:
        df = df[df["매입처"].isin(suppliers)]
    if reps:
        df = df[df["담당자"].isin(reps)]
    if min_s is not None:
        df = df[df["실매출금액"] >= min_s]
    if max_s is not None:
        df = df[df["실매출금액"] <= max_s]
    return df


def safe_records(df):
    records = df.to_dict(orient="records")
    cleaned = []
    for r in records:
        cleaned.append({k: (0 if isinstance(v, float) and math.isnan(v) else v) for k, v in r.items()})
    return cleaned


COMMON = dict(
    years=Query(default=[]),
    customers=Query(default=[]),
    manufacturers=Query(default=[]),
    suppliers=Query(default=[]),
    reps=Query(default=[]),
    min_sales=Query(default=None),
    max_sales=Query(default=None),
)


@app.get("/api/options")
def get_options():
    df = load_data()
    return {
        "years": sorted(df["연도"].unique().tolist()),
        "customers": sorted(df["매출처"].dropna().unique().tolist()),
        "manufacturers": sorted(df["제조사"].dropna().unique().tolist()),
        "suppliers": sorted(df["매입처"].dropna().unique().tolist()),
        "reps": sorted(df["담당자"].dropna().unique().tolist()),
        "min_sales": int(df["실매출금액"].min()),
        "max_sales": int(df["실매출금액"].max()),
    }


@app.get("/api/kpi")
def get_kpi(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    return {
        "total_sales": float(df["실매출금액"].sum()),
        "total_profit": float(df["실이익금액"].sum()),
        "avg_margin": float(df["실이익율"].mean()) if len(df) > 0 else 0.0,
        "total_qty": int(df["수량"].sum()),
        "total_count": int(len(df)),
    }


@app.get("/api/yearly")
def get_yearly(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    agg = df.groupby("연도").agg(
        sales=("실매출금액", "sum"),
        profit=("실이익금액", "sum"),
        avg_margin=("실이익율", "mean"),
        count=("수량", "count"),
        qty=("수량", "sum"),
    ).reset_index().fillna(0)
    return safe_records(agg)


@app.get("/api/customers")
def get_customers(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
    top_n: int = Query(default=10),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    by_sales = df.groupby("매출처")["실매출금액"].sum().nlargest(top_n).reset_index().fillna(0)
    by_profit = df.groupby("매출처")["실이익금액"].sum().nlargest(top_n).reset_index().fillna(0)
    top_names = by_sales["매출처"].tolist()
    by_year = df[df["매출처"].isin(top_names)].groupby(["매출처", "연도"])["실매출금액"].sum().reset_index().fillna(0)
    return {
        "by_sales": safe_records(by_sales),
        "by_profit": safe_records(by_profit),
        "by_year": safe_records(by_year),
    }


@app.get("/api/manufacturers")
def get_manufacturers(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
    top_n: int = Query(default=10),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    by_mfr = df.groupby("제조사")["실매출금액"].sum().nlargest(top_n).reset_index().fillna(0)
    by_sup = df.groupby("매입처")["실매출금액"].sum().nlargest(top_n).reset_index().fillna(0)
    pie = df.groupby("제조사")["실매출금액"].sum().nlargest(10).reset_index().fillna(0)
    return {
        "by_manufacturer": safe_records(by_mfr),
        "by_supplier": safe_records(by_sup),
        "pie": safe_records(pie),
    }


@app.get("/api/reps")
def get_reps(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
    top_n: int = Query(default=10),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    summary = df.groupby("담당자").agg(
        sales=("실매출금액", "sum"),
        profit=("실이익금액", "sum"),
        count=("수량", "count"),
        avg_margin=("실이익율", "mean"),
    ).reset_index().sort_values("sales", ascending=False).fillna(0)
    top_names = summary.head(top_n)["담당자"].tolist()
    by_year = df[df["담당자"].isin(top_names)].groupby(["담당자", "연도"])["실매출금액"].sum().reset_index().fillna(0)
    return {
        "summary": safe_records(summary),
        "top_n": safe_records(summary.head(top_n)),
        "by_year": safe_records(by_year),
    }


@app.get("/api/products")
def get_products(
    years: List[str] = Query(default=[]),
    customers: List[str] = Query(default=[]),
    manufacturers: List[str] = Query(default=[]),
    suppliers: List[str] = Query(default=[]),
    reps: List[str] = Query(default=[]),
    min_sales: Optional[float] = Query(default=None),
    max_sales: Optional[float] = Query(default=None),
    top_n: int = Query(default=10),
    search: Optional[str] = Query(default=None),
):
    df = filter_df(load_data(), years, customers, manufacturers, suppliers, reps, min_sales, max_sales)
    by_sales = df.groupby("제품명")["실매출금액"].sum().nlargest(top_n).reset_index().fillna(0)
    by_qty = df.groupby("제품명")["수량"].sum().nlargest(top_n).reset_index().fillna(0)
    margin = df["실이익율"].dropna()
    margin = margin[(margin > -100) & (margin < 100)]
    margin_avg = float(margin.mean()) if len(margin) > 0 else 0.0
    hist_data = []
    if len(margin) > 0:
        counts, edges = pd.cut(margin, bins=20, retbins=True)
        vc = counts.value_counts(sort=False)
        hist_data = [{"x": f"{edges[i]:.0f}~{edges[i+1]:.0f}", "y": int(c)} for i, c in enumerate(vc)]
    search_results = []
    if search:
        cols = [c for c in ["연도", "제품명", "규격", "제조사", "매출처", "담당자", "수량", "실매출금액", "실이익금액", "실이익율"] if c in df.columns]
        res = df[df["제품명"].str.contains(search, na=False)][cols].fillna("")
        search_results = safe_records(res)
    return {
        "by_sales": safe_records(by_sales),
        "by_qty": safe_records(by_qty),
        "margin_hist": hist_data,
        "margin_avg": margin_avg,
        "search": search_results,
    }


class EntryModel(BaseModel):
    year: str
    supplier: str
    manufacturer: str
    customer: str
    inv_loc: str = ""
    product: str
    spec: str = ""
    ins_code: str = ""
    ref_price: float = 0
    qty: int = 1
    unit_price: float = 0
    actual_unit_price: float = 0
    actual_unit_purch: float = 0
    rep: str


@app.post("/api/entry")
def add_entry(entry: EntryModel):
    filepath = FOLDER / f"{entry.year}.xlsx"
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"{entry.year}.xlsx 파일이 없습니다.")
    sales_amt = entry.unit_price * entry.qty
    act_sales = entry.actual_unit_price * entry.qty
    act_purch = entry.actual_unit_purch * entry.qty
    act_profit = act_sales - act_purch
    profit_rate = (act_profit / act_sales * 100) if act_sales > 0 else 0.0
    ref_sales = entry.ref_price * entry.qty
    wb = load_workbook(filepath)
    ws = wb.active
    seq_vals = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value is not None]
    next_seq = (max(seq_vals) + 1) if seq_vals else 1
    ws.append([
        next_seq, entry.supplier, entry.manufacturer, entry.customer,
        entry.inv_loc or entry.customer, entry.product, entry.spec,
        entry.ins_code, entry.ref_price, entry.qty, entry.unit_price,
        sales_amt, entry.actual_unit_price, act_sales, entry.actual_unit_purch,
        act_purch, act_profit, profit_rate, entry.rep, ref_sales,
    ])
    wb.save(filepath)
    return {"success": True, "seq": next_seq}


@app.get("/manifest.json")
def serve_manifest():
    return FileResponse(FOLDER / "static" / "manifest.json")


@app.get("/sw.js")
def serve_sw():
    return FileResponse(FOLDER / "static" / "sw.js")


app.mount("/", StaticFiles(directory=str(FOLDER / "static"), html=True), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=False)
